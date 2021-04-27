# -*- coding: utf-8 -*-
"""
Created on Mon Apr 26 20:52:25 2021

@author: Sergio
"""


import numpy as np

import astropy as astropy 

import pandas as pd

import seaborn as seaborn

import matplotlib.pyplot as plt

import matplotlib.ticker as ticker

import matplotlib.axis as ax

from astropy.table import Table, vstack    #Ojo importante aquí importar "Table" con mayuscula

from openpyxl import Workbook

#Leo datos (Objetos comunes de SHARKS y DES)
 

df=Table.read("fits/sharks_sgpe.fits", format="fits")  #Obj en sharks y des, con 5sigma


#Acoto objetos con signal/noise superior a 5 sigma.

#Para ello considero solo valores de magerr menores a 1.086/5

ks_mag = df["PETROMAG"]
ks_mag_error = df["PETROMAGERR"]

mask_5_sigma = (ks_mag_error <= 1.086/5)

#Filtro mi dataframe y lo renombro, ahora solo con objetos que considero validos (con signal/noise mayor que 5)

df = df[mask_5_sigma]

mag_ks = df["PETROMAG"]

mask_05 = (mag_ks<19.2)#&(mag_des_r <= 24.65)  #Si aplico las dos condiciones, se limita el numero de objetos (pierdo objetos)
mask_098 = (mag_ks>=19.2)#&(mag_des_r > 24.65)

df_098 = df[mask_098]
df_05 = df[mask_05]

#Empiezo con 0.5

classstat = df_05["CLASSSTAT"]

mask_galaxies = (classstat < 0.5)

mask_stars = (classstat >= 0.5)

df_galaxies = df_05[mask_galaxies]

df_stars = df_05[mask_stars]

#Ya tengo diferenciadas galaxias y estrellas

#Ahora puedo empezar a usar los dataframes que diferencian galaxias y estrellas con las máscaras numeradas, y sacar lo que me interesa.

#El sufijo "g" lo uso de ahora en adelante para considerar galaxias, y el sufijo "s" para estrellas (stars)

mag_ks_g = df_galaxies["PETROMAG"]

mask_0_g = (mag_ks_g >= 13.7)&(mag_ks_g < 14.2)
mask_1_g = (mag_ks_g >= 14.2)&(mag_ks_g < 14.7)
mask_2_g = (mag_ks_g >= 14.7)&(mag_ks_g < 15.2)
mask_3_g = (mag_ks_g >= 15.2)&(mag_ks_g < 15.7)
mask_4_g = (mag_ks_g >= 15.7)&(mag_ks_g < 16.2)
mask_5_g = (mag_ks_g >= 16.2)&(mag_ks_g < 16.7)
mask_6_g = (mag_ks_g >= 16.7)&(mag_ks_g < 17.2)
mask_7_g = (mag_ks_g >= 17.2)&(mag_ks_g < 17.7)
mask_8_g = (mag_ks_g >= 17.7)&(mag_ks_g < 18.2)
mask_9_g = (mag_ks_g >= 18.2)&(mag_ks_g < 18.7)
mask_10_g = (mag_ks_g >= 18.7)&(mag_ks_g < 19.2)
mask_11_g = (mag_ks_g >= 19.2)&(mag_ks_g < 19.7)
mask_12_g = (mag_ks_g >= 19.7)&(mag_ks_g < 20.2)
mask_13_g = (mag_ks_g >= 20.2)&(mag_ks_g < 20.7)
mask_14_g = (mag_ks_g >= 20.7)&(mag_ks_g < 21.2)
mask_15_g = (mag_ks_g >= 21.2)&(mag_ks_g < 21.7)
mask_16_g = (mag_ks_g >= 21.7)&(mag_ks_g < 22.2)
mask_17_g = (mag_ks_g >= 22.2)&(mag_ks_g < 22.7)

df_galaxies_0 = df_galaxies[mask_0_g]
df_galaxies_1 = df_galaxies[mask_1_g]
df_galaxies_2 = df_galaxies[mask_2_g]
df_galaxies_3 = df_galaxies[mask_3_g]
df_galaxies_4 = df_galaxies[mask_4_g]
df_galaxies_5 = df_galaxies[mask_5_g]
df_galaxies_6 = df_galaxies[mask_6_g]
df_galaxies_7 = df_galaxies[mask_7_g]
df_galaxies_8 = df_galaxies[mask_8_g]
df_galaxies_9 = df_galaxies[mask_9_g]
df_galaxies_10 = df_galaxies[mask_10_g]
df_galaxies_11 = df_galaxies[mask_11_g]
df_galaxies_12 = df_galaxies[mask_12_g]
df_galaxies_13 = df_galaxies[mask_13_g]
df_galaxies_14 = df_galaxies[mask_14_g]
df_galaxies_15 = df_galaxies[mask_15_g]
df_galaxies_16 = df_galaxies[mask_16_g]
df_galaxies_17 = df_galaxies[mask_17_g]

mag_ks_s = df_stars["PETROMAG"]

mask_0_s = (mag_ks_s >= 13.7)&(mag_ks_s < 14.2)
mask_1_s =  (mag_ks_s >= 14.2)&(mag_ks_s < 14.7)
mask_2_s = (mag_ks_s >= 14.7)&(mag_ks_s < 15.2)
mask_3_s = (mag_ks_s >= 15.2)&(mag_ks_s < 15.7)
mask_4_s = (mag_ks_s >= 15.7)&(mag_ks_s < 16.2)
mask_5_s = (mag_ks_s >= 16.2)&(mag_ks_s < 16.7)
mask_6_s = (mag_ks_s >= 16.7)&(mag_ks_s < 17.2)
mask_7_s = (mag_ks_s >= 17.2)&(mag_ks_s < 17.7)
mask_8_s = (mag_ks_s >= 17.7)&(mag_ks_s < 18.2)
mask_9_s = (mag_ks_s >= 18.2)&(mag_ks_s < 18.7)
mask_10_s = (mag_ks_s >= 18.7)&(mag_ks_s < 19.2)
mask_11_s = (mag_ks_s >= 19.2)&(mag_ks_s < 19.7)
mask_12_s = (mag_ks_s >= 19.7)&(mag_ks_s < 20.2)
mask_13_s = (mag_ks_s >= 20.2)&(mag_ks_s < 20.7)
mask_14_s = (mag_ks_s >= 20.7)&(mag_ks_s < 21.2)
mask_15_s = (mag_ks_s >= 21.2)&(mag_ks_s < 21.7)
mask_16_s = (mag_ks_s >= 21.7)&(mag_ks_s < 22.2)
mask_17_s = (mag_ks_s >= 22.2)&(mag_ks_s < 22.7)

df_stars_0 = df_stars[mask_0_s]
df_stars_1 = df_stars[mask_1_s]
df_stars_2 = df_stars[mask_2_s]
df_stars_3 = df_stars[mask_3_s]
df_stars_4 = df_stars[mask_4_s]
df_stars_5 = df_stars[mask_5_s]
df_stars_6 = df_stars[mask_6_s]
df_stars_7 = df_stars[mask_7_s]
df_stars_8 = df_stars[mask_8_s]
df_stars_9 = df_stars[mask_9_s]
df_stars_10 = df_stars[mask_10_s]
df_stars_11 = df_stars[mask_11_s]
df_stars_12 = df_stars[mask_12_s]
df_stars_13 = df_stars[mask_13_s]
df_stars_14 = df_stars[mask_14_s]
df_stars_15 = df_stars[mask_15_s]
df_stars_16 = df_stars[mask_16_s]
df_stars_17 = df_stars[mask_17_s]



#Ahora calculo el num de estrellas y galaxias, para cada rango.

galaxie_0 = len(df_galaxies_0)
galaxie_1 = len(df_galaxies_1)
galaxie_2 = len(df_galaxies_2)
galaxie_3 = len(df_galaxies_3)
galaxie_4 = len(df_galaxies_4)
galaxie_5 = len(df_galaxies_5)
galaxie_6 = len(df_galaxies_6)
galaxie_7 = len(df_galaxies_7)
galaxie_8 = len(df_galaxies_8)
galaxie_9 = len(df_galaxies_9)
galaxie_10 = len(df_galaxies_10)
galaxie_11 = len(df_galaxies_11)
galaxie_12 = len(df_galaxies_12)
galaxie_13 = len(df_galaxies_13)
galaxie_14 = len(df_galaxies_14)
galaxie_15 = len(df_galaxies_15)
galaxie_16 = len(df_galaxies_16)
galaxie_17 = len(df_galaxies_17)



galaxies_list_05 = np.array([galaxie_0, galaxie_1, galaxie_2, galaxie_3, galaxie_4, galaxie_5, galaxie_6, galaxie_7, galaxie_8, galaxie_9, galaxie_10, galaxie_11, galaxie_12, galaxie_13, galaxie_14, galaxie_15, galaxie_16, galaxie_17])



star_0 = len(df_stars_0)
star_1 = len(df_stars_1)
star_2 = len(df_stars_2)
star_3 = len(df_stars_3)
star_4 = len(df_stars_4)
star_5 = len(df_stars_5)
star_6 = len(df_stars_6)
star_7 = len(df_stars_7)
star_8 = len(df_stars_8)
star_9 = len(df_stars_9)
star_10 = len(df_stars_10)
star_11 = len(df_stars_11)
star_12 = len(df_stars_12)
star_13 = len(df_stars_13)
star_14 = len(df_stars_14)
star_15 = len(df_stars_15)
star_16 = len(df_stars_16)
star_17 = len(df_stars_17)


stars_list_05 = np.array([star_0, star_1, star_2, star_3, star_4, star_5, star_6, star_7, star_8, star_9, star_10, star_11, star_12, star_13, star_14, star_15, star_16, star_17])



#Sigo con 0.98


classstat = df_098["CLASSSTAT"]

mask_galaxies = (classstat < 0.975)    #usar 0.975

mask_stars = (classstat >= 0.975)

df_galaxies = df_098[mask_galaxies]

df_stars = df_098[mask_stars]

#Ya tengo diferenciadas galaxias y estrellas

#Ahora puedo empezar a usar los dataframes que diferencian galaxias y estrellas con las máscaras numeradas, y sacar lo que me interesa.

#El sufijo "g" lo uso de ahora en adelante para considerar galaxias, y el sufijo "s" para estrellas (stars)

mag_ks_g = df_galaxies["PETROMAG"]

mask_0_g = (mag_ks_g >= 13.7)&(mag_ks_g < 14.2)
mask_1_g = (mag_ks_g >= 14.2)&(mag_ks_g < 14.7)
mask_2_g = (mag_ks_g >= 14.7)&(mag_ks_g < 15.2)
mask_3_g = (mag_ks_g >= 15.2)&(mag_ks_g < 15.7)
mask_4_g = (mag_ks_g >= 15.7)&(mag_ks_g < 16.2)
mask_5_g = (mag_ks_g >= 16.2)&(mag_ks_g < 16.7)
mask_6_g = (mag_ks_g >= 16.7)&(mag_ks_g < 17.2)
mask_7_g = (mag_ks_g >= 17.2)&(mag_ks_g < 17.7)
mask_8_g = (mag_ks_g >= 17.7)&(mag_ks_g < 18.2)
mask_9_g = (mag_ks_g >= 18.2)&(mag_ks_g < 18.7)
mask_10_g = (mag_ks_g >= 18.7)&(mag_ks_g < 19.2)
mask_11_g = (mag_ks_g >= 19.2)&(mag_ks_g < 19.7)
mask_12_g = (mag_ks_g >= 19.7)&(mag_ks_g < 20.2)
mask_13_g = (mag_ks_g >= 20.2)&(mag_ks_g < 20.7)
mask_14_g = (mag_ks_g >= 20.7)&(mag_ks_g < 21.2)
mask_15_g = (mag_ks_g >= 21.2)&(mag_ks_g < 21.7)
mask_16_g = (mag_ks_g >= 21.7)&(mag_ks_g < 22.2)
mask_17_g = (mag_ks_g >= 22.2)&(mag_ks_g < 22.7)

df_galaxies_0 = df_galaxies[mask_0_g]
df_galaxies_1 = df_galaxies[mask_1_g]
df_galaxies_2 = df_galaxies[mask_2_g]
df_galaxies_3 = df_galaxies[mask_3_g]
df_galaxies_4 = df_galaxies[mask_4_g]
df_galaxies_5 = df_galaxies[mask_5_g]
df_galaxies_6 = df_galaxies[mask_6_g]
df_galaxies_7 = df_galaxies[mask_7_g]
df_galaxies_8 = df_galaxies[mask_8_g]
df_galaxies_9 = df_galaxies[mask_9_g]
df_galaxies_10 = df_galaxies[mask_10_g]
df_galaxies_11 = df_galaxies[mask_11_g]
df_galaxies_12 = df_galaxies[mask_12_g]
df_galaxies_13 = df_galaxies[mask_13_g]
df_galaxies_14 = df_galaxies[mask_14_g]
df_galaxies_15 = df_galaxies[mask_15_g]
df_galaxies_16 = df_galaxies[mask_16_g]
df_galaxies_17 = df_galaxies[mask_17_g]

mag_ks_s = df_stars["PETROMAG"]

mask_0_s = (mag_ks_s >= 13.7)&(mag_ks_s < 14.2)
mask_1_s =  (mag_ks_s >= 14.2)&(mag_ks_s < 14.7)
mask_2_s = (mag_ks_s >= 14.7)&(mag_ks_s < 15.2)
mask_3_s = (mag_ks_s >= 15.2)&(mag_ks_s < 15.7)
mask_4_s = (mag_ks_s >= 15.7)&(mag_ks_s < 16.2)
mask_5_s = (mag_ks_s >= 16.2)&(mag_ks_s < 16.7)
mask_6_s = (mag_ks_s >= 16.7)&(mag_ks_s < 17.2)
mask_7_s = (mag_ks_s >= 17.2)&(mag_ks_s < 17.7)
mask_8_s = (mag_ks_s >= 17.7)&(mag_ks_s < 18.2)
mask_9_s = (mag_ks_s >= 18.2)&(mag_ks_s < 18.7)
mask_10_s = (mag_ks_s >= 18.7)&(mag_ks_s < 19.2)
mask_11_s = (mag_ks_s >= 19.2)&(mag_ks_s < 19.7)
mask_12_s = (mag_ks_s >= 19.7)&(mag_ks_s < 20.2)
mask_13_s = (mag_ks_s >= 20.2)&(mag_ks_s < 20.7)
mask_14_s = (mag_ks_s >= 20.7)&(mag_ks_s < 21.2)
mask_15_s = (mag_ks_s >= 21.2)&(mag_ks_s < 21.7)
mask_16_s = (mag_ks_s >= 21.7)&(mag_ks_s < 22.2)
mask_17_s = (mag_ks_s >= 22.2)&(mag_ks_s < 22.7)

df_stars_0 = df_stars[mask_0_s]
df_stars_1 = df_stars[mask_1_s]
df_stars_2 = df_stars[mask_2_s]
df_stars_3 = df_stars[mask_3_s]
df_stars_4 = df_stars[mask_4_s]
df_stars_5 = df_stars[mask_5_s]
df_stars_6 = df_stars[mask_6_s]
df_stars_7 = df_stars[mask_7_s]
df_stars_8 = df_stars[mask_8_s]
df_stars_9 = df_stars[mask_9_s]
df_stars_10 = df_stars[mask_10_s]
df_stars_11 = df_stars[mask_11_s]
df_stars_12 = df_stars[mask_12_s]
df_stars_13 = df_stars[mask_13_s]
df_stars_14 = df_stars[mask_14_s]
df_stars_15 = df_stars[mask_15_s]
df_stars_16 = df_stars[mask_16_s]
df_stars_17 = df_stars[mask_17_s]



#Ahora calculo el num de estrellas y galaxias, para cada rango.

galaxie_0 = len(df_galaxies_0)
galaxie_1 = len(df_galaxies_1)
galaxie_2 = len(df_galaxies_2)
galaxie_3 = len(df_galaxies_3)
galaxie_4 = len(df_galaxies_4)
galaxie_5 = len(df_galaxies_5)
galaxie_6 = len(df_galaxies_6)
galaxie_7 = len(df_galaxies_7)
galaxie_8 = len(df_galaxies_8)
galaxie_9 = len(df_galaxies_9)
galaxie_10 = len(df_galaxies_10)
galaxie_11 = len(df_galaxies_11)
galaxie_12 = len(df_galaxies_12)
galaxie_13 = len(df_galaxies_13)
galaxie_14 = len(df_galaxies_14)
galaxie_15 = len(df_galaxies_15)
galaxie_16 = len(df_galaxies_16)
galaxie_17 = len(df_galaxies_17)



galaxies_list_098 = np.array([galaxie_0, galaxie_1, galaxie_2, galaxie_3, galaxie_4, galaxie_5, galaxie_6, galaxie_7, galaxie_8, galaxie_9, galaxie_10, galaxie_11, galaxie_12, galaxie_13, galaxie_14, galaxie_15, galaxie_16, galaxie_17])



star_0 = len(df_stars_0)
star_1 = len(df_stars_1)
star_2 = len(df_stars_2)
star_3 = len(df_stars_3)
star_4 = len(df_stars_4)
star_5 = len(df_stars_5)
star_6 = len(df_stars_6)
star_7 = len(df_stars_7)
star_8 = len(df_stars_8)
star_9 = len(df_stars_9)
star_10 = len(df_stars_10)
star_11 = len(df_stars_11)
star_12 = len(df_stars_12)
star_13 = len(df_stars_13)
star_14 = len(df_stars_14)
star_15 = len(df_stars_15)
star_16 = len(df_stars_16)
star_17 = len(df_stars_17)


stars_list_098 = np.array([star_0, star_1, star_2, star_3, star_4, star_5, star_6, star_7, star_8, star_9, star_10, star_11, star_12, star_13, star_14, star_15, star_16, star_17])


stars_total = stars_list_05 + stars_list_098
galaxies_total = galaxies_list_05 + galaxies_list_098

area_sharks = 7.23 #revisar, este dato no lo recuero y perdí el correo que me decía Aurelio, es en grados de arco

N_stars = stars_total/area_sharks
N_galaxies = galaxies_total/area_sharks

N_stars_err = np.sqrt(stars_total)/area_sharks
N_galaxies_err = np.sqrt(galaxies_total)/area_sharks

ks_sharks_vega = [12, 12.5, 13, 13.5, 14, 14.5, 15, 15.5, 16, 16.5, 17, 17.5, 18, 18.5, 19, 19.5, 20, 20.5]

ks_sharks_AB = []

for i in range(len(ks_sharks_vega)):
    ks_sharks_AB.append(ks_sharks_vega[i]+1.83)


ks_Daddi_vega = [12, 12.5, 13, 13.5, 14, 14.5, 15, 15.5, 16, 16.5, 17, 17.5, 18, 18.5, 19]

ks_Daddi_AB = []

for i in range(len(ks_Daddi_vega)):
    ks_Daddi_AB.append(ks_Daddi_vega[i]+1.83)
    

stars_Daddi = np.array([4, 7, 9, 17, 21, 38, 37, 62, 73, 88, 128, 127, 144, 185, 84])  #El intervalo 18.7-18.8 no lo representa

galaxies_Daddi = np.array([0, 0, 0, 0, 4, 6, 16, 30, 74, 100, 178, 372, 633, 892, 628])

area_Daddi_1 = 701/3600   #En deg^2 (paso de arcmin^2 a deg^2)

N_stars_Daddi = stars_Daddi[0:14]/area_Daddi_1           #De 0 al 13
N_galaxies_Daddi = galaxies_Daddi[0:14]/area_Daddi_1


area_Daddi_2 = 447.5/3600


N_stars_Daddi = np.append(N_stars_Daddi, stars_Daddi[14]/area_Daddi_2)
N_galaxies_Daddi = np.append(N_galaxies_Daddi, galaxies_Daddi[14]/area_Daddi_2)


#Log(N) viene dividido siempre por 2.33, que es 0.5+1.83. Esto es un poco raro, pero así parece que salen los números del artículo de Daddi, 
#aunque no tiene mucho sentido. En el TFG puedo omitir esto. 

#Añado el error poissoniano también de mis muestras, que es la raíz del número de cuentas.


plt.figure()
#plt.plot(ks_Daddi_AB,np.log(N_stars_Daddi), "x" ,label="Stars in Daddi´s article")         
#plt.plot(ks_Daddi_AB,np.log(N_galaxies_Daddi), "s" ,label="Galaxies in Daddi´s article")  
plt.errorbar(ks_sharks_AB,np.log(N_stars), np.log(N_stars_err), elinewidth=0.5, barsabove= True, color = "cornflowerblue", marker = "*" , linestyle="None", label="Stars in Sharks")
#plt.plot(ks_sharks_AB,np.log(N_stars), "*" ,label="Stars in Sharks")      
plt.errorbar(ks_sharks_AB,np.log(N_galaxies), np.log(N_galaxies_err), elinewidth=0.5, barsabove= True , color = "royalblue", marker = "." , linestyle="None", label="Galaxies in Sharks")         
#plt.plot(ks_sharks_AB,np.log(N_galaxies), "." ,label="Galaxies in Sharks")
plt.xlabel("Ks", fontsize = 12)
plt.ylabel(r"$log N ~ (objects/deg^2)$", fontsize=12)
plt.xticks([13, 15, 17, 19, 21, 23 ])
#plt.yticks([])   #Conviene añadir un yticks para quitarse los logaritmos negativos/menores a uno.
plt.title("Stars and galaxies number of counts comparison")
plt.grid(linestyle="--", linewidth=0.2)
plt.legend()
plt.savefig("Stars_galaxies_of_sharks_with_errors.png")



plt.figure()
plt.plot(ks_Daddi_AB,np.log(N_stars_Daddi),  "*" , color = "sandybrown", label="Stars in Daddi´s article")         
plt.plot(ks_Daddi_AB,np.log(N_galaxies_Daddi), "." ,color = "orangered", label="Galaxies in Daddi´s article")  
#plt.errorbar(ks_sharks_AB,np.log(N_stars), np.log(N_stars_err), elinewidth=0.5, barsabove= True , marker = "*" , linestyle="None", label="Stars in Sharks")
plt.plot(ks_sharks_AB,np.log(N_stars), "*" ,color = "cornflowerblue", label="Stars in Sharks")      
#plt.errorbar(ks_sharks_AB,np.log(N_galaxies), np.log(N_galaxies_err), elinewidth=0.5, barsabove= True , marker = "." , linestyle="None", label="Galaxies in Sharks")         
plt.plot(ks_sharks_AB,np.log(N_galaxies), "." , color = "royalblue", label="Galaxies in Sharks")
plt.xlabel("Ks", fontsize = 12)
plt.ylabel(r"$log N ~ (objects/deg^2)$", fontsize=12)
plt.xticks([13, 15, 17, 19, 21, 23 ])
#plt.yticks([])   #Conviene añadir un yticks para quitarse los logaritmos negativos/menores a uno.
plt.title("Stars and galaxies number of counts comparison")
plt.grid(linestyle="--", linewidth=0.2)
plt.legend()
plt.savefig("Stars_galaxies_comparison_with_Daddi_RESTRICTED.png")







wb = Workbook()
ruta = 'salida_ALL_restriccion.xlsx'

hoja = wb.active
hoja.title = "ALL"

fila = 1 #Fila donde empezamos
col_stars = 1 #Columna donde guardo los valores
col_galaxies = 2

for stars, galaxies in zip(stars_total, galaxies_total):  #Aquí pongo en "in" el valor que quiero sacar como columna
    hoja.cell(column=col_stars, row=fila, value=stars)
    hoja.cell(column=col_galaxies, row=fila, value=galaxies)
    fila+=1

wb.save(filename = ruta)




