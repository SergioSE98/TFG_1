# -*- coding: utf-8 -*-
"""
Created on Sat Apr 17 14:32:13 2021

@author: Sergio
"""

import numpy as np

import astropy as astropy 

import pandas as pd

import seaborn as seaborn

import matplotlib.pyplot as plt

from astropy.table import Table, vstack    #Ojo importante aquí importar "Table" con mayuscula

from openpyxl import Workbook

#Leo datos (Objetos comunes de SHARKS y DES)
 

df_1=Table.read("fits/sharks_and_des_sig_noise_5_lite.fits", format="fits")  #Obj en sharks y des, con 5sigma

df_2 = Table.read("fits/sharks_only_not_des_lite.fits", format="fits") #Obj solo en Sharks, muestra ya limpiada de difracciones y grandes errores

#El dataframe 1 ya está limpio de sig/noise menor a 5, el 2 lo limpio ahora.

ks_mag_error = df_2["PETROMAGERR"]
mask_5_sigma = (ks_mag_error <= 1.086/5)

#Filtro mi dataframe y lo renombro, ahora solo con objetos que considero validos (con signal/noise mayor que 5)

df_2 = df_2[mask_5_sigma]

#Ahora diferencio EROS en el dataframe1, los del taframe 2 son todo EROS.

mag_sharks = df_1["PETROMAG"]   #Consultar si debería cambiar esto por magnitud de otra apertura.
mag_err_sharks = df_1["PETROMAGERR"]
mag_des_i = df_1["MAG_AUTO_I_DERED"]
mag_des_g = df_1["MAG_AUTO_G_DERED"]
mag_des_r = df_1["MAG_AUTO_R_DERED"]

eros_mask = ((mag_des_r - mag_sharks) > 4.32)   #En primer lugar filtro los EROS del df con obj en sharks y des, los de solo sharks es pq directamente ya se pueden tratar como eros (no se detectan en optico)

df_eros = df_1[eros_mask]   #Cambio mi df y lo acoto a EROS


#Esto anterior es para obtener info adicional, directamente puedo juntar los df df_2 y df_eros para sacar el dataframe total de eros (de sharks y des, y los exclusivos de sharks)


df = vstack([df_eros,df_2])  #Este es el dataframe que recoge todos los eros


#Utilizo los EROS que tienen detección en r.

mag_ks = df["PETROMAG"]   #Consultar si debería cambiar esto por magnitud de otra apertura.
mag_err_sharks = df["PETROMAGERR"]
mag_des_i = df["MAG_AUTO_I_DERED"]
mag_des_g = df["MAG_AUTO_G_DERED"]
mag_des_r = df["MAG_AUTO_R_DERED"]
classstat = df["CLASSSTAT"]



#El límite de detección de Sharks es 22.7 en banda Ks (AB, 5 sigma, lo que tengo yo), no hace falta acotarlo, las medidas llegan approx hasta ahí
#Voy a acotar distintos rangos de magnitud y ver cuántos objetos detecto en cada uno. 

#La mínima magnitud la tomo como en el artículo de daddi, que es 11.7+1.83 = 13.5 (La tomo = 13.7, mucha presencia de contaminación por difracción)

#Establezco entonces rangos 13.7-14.2,    ...,   22.2-22.7

#Creo también un array con el valor intermedio de cada intervalo 






#Para la clasificación de objetos en estrella/galaxia, la tomo valida para ks menor igual que 20  (sería 19.33 realmente), y r menor igual que 24.65 (no considero esto)

#Los objetos con mag mayor son muy tenues y pueden considerarse directamente galaxias, con una pequeña contaminación de enanas marrones (por ser tan tenues)

#Realizo entonces el acotado, ya no pondré lo de eros_r, estoy trabajando solo con esos.

#Al final no utilizo sesgo, aplico la diferenciación con classstat a todos los objetos, si quiero usar diferenciacion usar sig filas del codigo

mask_classstat = (mag_ks<21.2)#&(mag_des_r <= 24.65)  #Si aplico las dos condiciones, se limita el numero de objetos (pierdo objetos)
mask_all_galaxies = (mag_ks>=21.2)#&(mag_des_r > 24.65)

df_all_galaxies = df[mask_all_galaxies]
df = df[mask_classstat]
classstat = df["CLASSSTAT"]



#Diferencio el df total entre galaxias y estrellas

mask_galaxies = (classstat < 0.5)

mask_stars = (classstat >= 0.5)

df_galaxies = df[mask_galaxies]

df_stars = df[mask_stars]

#Ya tengo diferenciadas galaxias y estrellas

#Ahora puedo empezar a usar los dataframes que diferencian galaxias y estrellas con las máscaras numeradas, y sacar lo que me interesa.

#El sufijo "g" lo uso de ahora en adelante para considerar galaxias, y el sufijo "s" para estrellas (stars)

mag_ks_g = df_galaxies["PETROMAG"]

mask_0_g = (mag_ks_g >= 13.7)&(mag_ks_g < 14.2)
mask_1_g = (mag_ks_g >= 14.2)&(mag_ks_g < 14.7)
mask_2_g = (mag_ks_g >= 14.7)&(mag_ks_g < 15.2)
mask_3_g = (mag_ks_g >= 15.2)&(mag_ks_g < 15.7)
mask_4_g = (mag_ks_g >= 15.7)&(mag_ks_g < 16.2)
mask_5_g = (mag_ks_g >= 16.2)&(mag_ks_g < 16.7)
mask_6_g = (mag_ks_g >= 16.7)&(mag_ks_g < 17.2)
mask_7_g = (mag_ks_g >= 17.2)&(mag_ks_g < 17.7)
mask_8_g = (mag_ks_g >= 17.7)&(mag_ks_g < 18.2)
mask_9_g = (mag_ks_g >= 18.2)&(mag_ks_g < 18.7)
mask_10_g = (mag_ks_g >= 18.7)&(mag_ks_g < 19.2)
mask_11_g = (mag_ks_g >= 19.2)&(mag_ks_g < 19.7)
mask_12_g = (mag_ks_g >= 19.7)&(mag_ks_g < 20.2)
mask_13_g = (mag_ks_g >= 20.2)&(mag_ks_g < 20.7)
mask_14_g = (mag_ks_g >= 20.7)&(mag_ks_g < 21.2)
mask_15_g = (mag_ks_g >= 21.2)&(mag_ks_g < 21.7)
mask_16_g = (mag_ks_g >= 21.7)&(mag_ks_g < 22.2)
mask_17_g = (mag_ks_g >= 22.2)&(mag_ks_g < 22.7)

df_galaxies_0 = df_galaxies[mask_0_g]
df_galaxies_1 = df_galaxies[mask_1_g]
df_galaxies_2 = df_galaxies[mask_2_g]
df_galaxies_3 = df_galaxies[mask_3_g]
df_galaxies_4 = df_galaxies[mask_4_g]
df_galaxies_5 = df_galaxies[mask_5_g]
df_galaxies_6 = df_galaxies[mask_6_g]
df_galaxies_7 = df_galaxies[mask_7_g]
df_galaxies_8 = df_galaxies[mask_8_g]
df_galaxies_9 = df_galaxies[mask_9_g]
df_galaxies_10 = df_galaxies[mask_10_g]
df_galaxies_11 = df_galaxies[mask_11_g]
df_galaxies_12 = df_galaxies[mask_12_g]
df_galaxies_13 = df_galaxies[mask_13_g]
df_galaxies_14 = df_galaxies[mask_14_g]
df_galaxies_15 = df_galaxies[mask_15_g]
df_galaxies_16 = df_galaxies[mask_16_g]
df_galaxies_17 = df_galaxies[mask_17_g]

mag_ks_s = df_stars["PETROMAG"]

mask_0_s = (mag_ks_s >= 13.7)&(mag_ks_s < 14.2)
mask_1_s =  (mag_ks_s >= 14.2)&(mag_ks_s < 14.7)
mask_2_s = (mag_ks_s >= 14.7)&(mag_ks_s < 15.2)
mask_3_s = (mag_ks_s >= 15.2)&(mag_ks_s < 15.7)
mask_4_s = (mag_ks_s >= 15.7)&(mag_ks_s < 16.2)
mask_5_s = (mag_ks_s >= 16.2)&(mag_ks_s < 16.7)
mask_6_s = (mag_ks_s >= 16.7)&(mag_ks_s < 17.2)
mask_7_s = (mag_ks_s >= 17.2)&(mag_ks_s < 17.7)
mask_8_s = (mag_ks_s >= 17.7)&(mag_ks_s < 18.2)
mask_9_s = (mag_ks_s >= 18.2)&(mag_ks_s < 18.7)
mask_10_s = (mag_ks_s >= 18.7)&(mag_ks_s < 19.2)
mask_11_s = (mag_ks_s >= 19.2)&(mag_ks_s < 19.7)
mask_12_s = (mag_ks_s >= 19.7)&(mag_ks_s < 20.2)
mask_13_s = (mag_ks_s >= 20.2)&(mag_ks_s < 20.7)
mask_14_s = (mag_ks_s >= 20.7)&(mag_ks_s < 21.2)
mask_15_s = (mag_ks_s >= 21.2)&(mag_ks_s < 21.7)
mask_16_s = (mag_ks_s >= 21.7)&(mag_ks_s < 22.2)
mask_17_s = (mag_ks_s >= 22.2)&(mag_ks_s < 22.7)

df_stars_0 = df_stars[mask_0_s]
df_stars_1 = df_stars[mask_1_s]
df_stars_2 = df_stars[mask_2_s]
df_stars_3 = df_stars[mask_3_s]
df_stars_4 = df_stars[mask_4_s]
df_stars_5 = df_stars[mask_5_s]
df_stars_6 = df_stars[mask_6_s]
df_stars_7 = df_stars[mask_7_s]
df_stars_8 = df_stars[mask_8_s]
df_stars_9 = df_stars[mask_9_s]
df_stars_10 = df_stars[mask_10_s]
df_stars_11 = df_stars[mask_11_s]
df_stars_12 = df_stars[mask_12_s]
df_stars_13 = df_stars[mask_13_s]
df_stars_14 = df_stars[mask_14_s]
df_stars_15 = df_stars[mask_15_s]
df_stars_16 = df_stars[mask_16_s]
df_stars_17 = df_stars[mask_17_s]

mag_ks_all_g = df_all_galaxies["PETROMAG"]

mask_0_all_g = (mag_ks_all_g >= 13.7)&(mag_ks_all_g < 14.2)
mask_1_all_g =  (mag_ks_all_g >= 14.2)&(mag_ks_all_g < 14.7)
mask_2_all_g = (mag_ks_all_g >= 14.7)&(mag_ks_all_g < 15.2)
mask_3_all_g = (mag_ks_all_g >= 15.2)&(mag_ks_all_g < 15.7)
mask_4_all_g = (mag_ks_all_g >= 15.7)&(mag_ks_all_g < 16.2)
mask_5_all_g = (mag_ks_all_g >= 16.2)&(mag_ks_all_g < 16.7)
mask_6_all_g = (mag_ks_all_g >= 16.7)&(mag_ks_all_g < 17.2)
mask_7_all_g = (mag_ks_all_g >= 17.2)&(mag_ks_all_g < 17.7)
mask_8_all_g = (mag_ks_all_g >= 17.7)&(mag_ks_all_g < 18.2)
mask_9_all_g = (mag_ks_all_g >= 18.2)&(mag_ks_all_g < 18.7)
mask_10_all_g = (mag_ks_all_g >= 18.7)&(mag_ks_all_g < 19.2)
mask_11_all_g = (mag_ks_all_g >= 19.2)&(mag_ks_all_g < 19.7)
mask_12_all_g = (mag_ks_all_g >= 19.7)&(mag_ks_all_g < 20.2)
mask_13_all_g = (mag_ks_all_g >= 20.2)&(mag_ks_all_g < 20.7)
mask_14_all_g = (mag_ks_all_g >= 20.7)&(mag_ks_all_g < 21.2)
mask_15_all_g = (mag_ks_all_g >= 21.2)&(mag_ks_all_g < 21.7)
mask_16_all_g = (mag_ks_all_g >= 21.7)&(mag_ks_all_g < 22.2)
mask_17_all_g = (mag_ks_all_g >= 22.2)&(mag_ks_all_g < 22.7)

df_all_galaxies_0 = df_all_galaxies[mask_0_all_g]
df_all_galaxies_1 = df_all_galaxies[mask_1_all_g]
df_all_galaxies_2 = df_all_galaxies[mask_2_all_g]
df_all_galaxies_3 = df_all_galaxies[mask_3_all_g]
df_all_galaxies_4 = df_all_galaxies[mask_4_all_g]
df_all_galaxies_5 = df_all_galaxies[mask_5_all_g]
df_all_galaxies_6 = df_all_galaxies[mask_6_all_g]
df_all_galaxies_7 = df_all_galaxies[mask_7_all_g]
df_all_galaxies_8 = df_all_galaxies[mask_8_all_g]
df_all_galaxies_9 = df_all_galaxies[mask_9_all_g]
df_all_galaxies_10 = df_all_galaxies[mask_10_all_g]
df_all_galaxies_11 = df_all_galaxies[mask_11_all_g]
df_all_galaxies_12 = df_all_galaxies[mask_12_all_g]
df_all_galaxies_13 = df_all_galaxies[mask_13_all_g]
df_all_galaxies_14 = df_all_galaxies[mask_14_all_g]
df_all_galaxies_15 = df_all_galaxies[mask_15_all_g]
df_all_galaxies_16 = df_all_galaxies[mask_16_all_g]
df_all_galaxies_17 = df_all_galaxies[mask_17_all_g]


#Ahora calculo el num de estrellas y galaxias, para cada rango.

galaxie_0 = len(df_galaxies_0)
galaxie_1 = len(df_galaxies_1)
galaxie_2 = len(df_galaxies_2)
galaxie_3 = len(df_galaxies_3)
galaxie_4 = len(df_galaxies_4)
galaxie_5 = len(df_galaxies_5)
galaxie_6 = len(df_galaxies_6)
galaxie_7 = len(df_galaxies_7)
galaxie_8 = len(df_galaxies_8)
galaxie_9 = len(df_galaxies_9)
galaxie_10 = len(df_galaxies_10)
galaxie_11 = len(df_galaxies_11)
galaxie_12 = len(df_galaxies_12)
galaxie_13 = len(df_galaxies_13)
galaxie_14 = len(df_galaxies_14)
galaxie_15 = len(df_galaxies_15)
galaxie_16 = len(df_galaxies_16)
galaxie_17 = len(df_galaxies_17)



galaxies_list = ([galaxie_0, galaxie_1, galaxie_2, galaxie_3, galaxie_4, galaxie_5, galaxie_6, galaxie_7, galaxie_8, galaxie_9, galaxie_10, galaxie_11, galaxie_12, galaxie_13, galaxie_14, galaxie_15, galaxie_16, galaxie_17])



star_0 = len(df_stars_0)
star_1 = len(df_stars_1)
star_2 = len(df_stars_2)
star_3 = len(df_stars_3)
star_4 = len(df_stars_4)
star_5 = len(df_stars_5)
star_6 = len(df_stars_6)
star_7 = len(df_stars_7)
star_8 = len(df_stars_8)
star_9 = len(df_stars_9)
star_10 = len(df_stars_10)
star_11 = len(df_stars_11)
star_12 = len(df_stars_12)
star_13 = len(df_stars_13)
star_14 = len(df_stars_14)
star_15 = len(df_stars_15)
star_16 = len(df_stars_16)
star_17 = len(df_stars_17)


stars_list = ([star_0, star_1, star_2, star_3, star_4, star_5, star_6, star_7, star_8, star_9, star_10, star_11, star_12, star_13, star_14, star_15, star_16, star_17])


all_galaxie_0 = len(df_all_galaxies_0)
all_galaxie_1 = len(df_all_galaxies_1)
all_galaxie_2 = len(df_all_galaxies_2)
all_galaxie_3 = len(df_all_galaxies_3)
all_galaxie_4 = len(df_all_galaxies_4)
all_galaxie_5 = len(df_all_galaxies_5)
all_galaxie_6 = len(df_all_galaxies_6)
all_galaxie_7 = len(df_all_galaxies_7)
all_galaxie_8 = len(df_all_galaxies_8)
all_galaxie_9 = len(df_all_galaxies_9)
all_galaxie_10 = len(df_all_galaxies_10)
all_galaxie_11 = len(df_all_galaxies_11)
all_galaxie_12 = len(df_all_galaxies_12)
all_galaxie_13 = len(df_all_galaxies_13)
all_galaxie_14 = len(df_all_galaxies_14)
all_galaxie_15 = len(df_all_galaxies_15)
all_galaxie_16 = len(df_all_galaxies_16)
all_galaxie_17 = len(df_all_galaxies_17)


all_galaxies_list = ([all_galaxie_0, all_galaxie_1, all_galaxie_2, all_galaxie_3, all_galaxie_4, all_galaxie_5, all_galaxie_6, all_galaxie_7, all_galaxie_8, all_galaxie_9, all_galaxie_10, all_galaxie_11, all_galaxie_12, all_galaxie_13, all_galaxie_14, all_galaxie_15, all_galaxie_16, all_galaxie_17])


total_galaxie_0 = galaxie_0 + all_galaxie_0
total_galaxie_1 = galaxie_1 + all_galaxie_1
total_galaxie_2 = galaxie_2 + all_galaxie_2
total_galaxie_3 = galaxie_3 + all_galaxie_3
total_galaxie_4 = galaxie_4 + all_galaxie_4
total_galaxie_5 = galaxie_5 + all_galaxie_5
total_galaxie_6 = galaxie_6 + all_galaxie_6
total_galaxie_7 = galaxie_7 + all_galaxie_7
total_galaxie_8 = galaxie_8 + all_galaxie_8
total_galaxie_9 = galaxie_9 + all_galaxie_9
total_galaxie_10 = galaxie_10 + all_galaxie_10
total_galaxie_11 = galaxie_11 + all_galaxie_11
total_galaxie_12 = galaxie_12 + all_galaxie_12
total_galaxie_13 = galaxie_13 + all_galaxie_13
total_galaxie_14 = galaxie_14 + all_galaxie_14
total_galaxie_15 = galaxie_15 + all_galaxie_15
total_galaxie_16 = galaxie_16 + all_galaxie_16
total_galaxie_17 = galaxie_17 + all_galaxie_17


total_galaxies_list = ([total_galaxie_0, total_galaxie_1, total_galaxie_2, total_galaxie_3, total_galaxie_4, total_galaxie_5, total_galaxie_6, total_galaxie_7, total_galaxie_8, total_galaxie_9, total_galaxie_10, total_galaxie_11, total_galaxie_12, total_galaxie_13, total_galaxie_14, total_galaxie_15, total_galaxie_16, total_galaxie_17])



ks_mitad_intervalos = ([14, 14.5, 15, 15.5, 16, 16.5, 17, 17.5, 18, 18.5, 19, 19.5, 20, 20.5, 21, 21.5, 22, 22.5]) #Esto lo represento en eje x
#Plots sencillitos, hago los más complejos luego


#He aplicado las mask a todos los intervalos, pero la realidad es que la lista de all_galaxies tendrá valores no nulos solo para índice mayor al 14,
#Y la lista de stars y galaxies solo no nulos para índice menor al 14








#Plots del informe

area = 7.23 #revisar, este dato no lo recuero y perdí el correo que me decía Aurelio, es en grados de arco
area_min = 7.23*3600


N_stars = list(map(lambda x: x/area, stars_list))
N_galaxies = list(map(lambda x: x/area, total_galaxies_list))

N_stars_mins = list(map(lambda x: x/area_min, stars_list))
N_galaxies_mins = list(map(lambda x: x/area_min, total_galaxies_list))


#xticks = np.arange(min(ks_mitad_intervalos),max(ks_mitad_intervalos),0.5 )

plt.figure()
plt.plot(ks_mitad_intervalos,np.log(N_stars), "*" ,label="Stars")
plt.plot(ks_mitad_intervalos,np.log(N_galaxies), "." ,label="Galaxies")
#plt.yscale("log")
plt.xlabel("Ks")
plt.ylabel(r"$log N ~ (objects/deg^2/0.5 mag)$")
#plt.ylim(0,10)
#plt.xlim(13.45,22.95)
plt.xticks(ks_mitad_intervalos)
plt.title("Stars and galaxies number of counts comparison")
plt.legend()


plt.figure()
plt.plot(ks_mitad_intervalos,np.log(N_galaxies), "." ,label="Galaxies")
plt.xlabel("Ks")
plt.ylabel(r"$log N ~ (objects/deg^2/0.5 mag)$")
plt.title("Galaxies number of counts")
plt.legend()
#Restrinjo más, considero menos magnitudes bajas


#Puedo también calcular la fracción de galaxias en cada sesgo respecto del total de galaxias, que saco de ALL.py, teniendo en cuenta que
#debo sumar las estrellas y galaxias de los últimos Ks, por encima de 21.2, y comparar entonces sí. 



"""


wb = Workbook()
ruta = 'salida_EROS.xlsx'


hoja1 = wb.active
hoja1.title = "hoja"
fila = 1 #Fila donde empezamos
col_stars = 1 #Columna donde guardo los valores
col_galaxies = 2
col_n_stars = 3
col_n_galaxies = 4
col_n_min_stars = 5
col_n_min_galaxies = 6

for stars, galaxies, n_stars, n_galaxies, n_min_stars, n_min_galaxies in zip(stars_list, total_galaxies_list, N_stars, N_galaxies, N_stars_mins, N_galaxies_mins):  #Aquí pongo en "in" el valor que quiero sacar como columna
    hoja1.cell(column=col_stars, row=fila, value=stars)
    hoja1.cell(column=col_galaxies, row=fila, value=galaxies)
    hoja1.cell(column=col_n_stars, row=fila, value=n_stars)
    hoja1.cell(column=col_n_galaxies, row=fila, value=n_galaxies)
    hoja1.cell(column=col_n_min_stars, row=fila, value=n_min_stars)
    hoja1.cell(column=col_n_min_galaxies, row=fila, value=n_min_galaxies)
    fila+=1
    


wb.save(filename = ruta)
"""