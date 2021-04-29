# -*- coding: utf-8 -*-
"""
Created on Sat Apr 17 14:32:13 2021

@author: Sergio
"""

import numpy as np

import astropy as astropy 

import pandas as pd

import seaborn as seaborn

import matplotlib.pyplot as plt

from astropy.table import Table, vstack    #Ojo importante aquí importar "Table" con mayuscula

from openpyxl import Workbook

#Leo datos (Objetos comunes de SHARKS y DES)
 

df=Table.read("fits/all_EROs.fits", format="fits")  #Obj en sharks y des, con 5sigma



mag_ks = df["PETROMAG"]

mask_05 = (mag_ks<19.2)#&(mag_des_r <= 24.65)  #Si aplico las dos condiciones, se limita el numero de objetos (pierdo objetos)
mask_098 = (mag_ks>=19.2)#&(mag_des_r > 24.65)

df_098 = df[mask_098]
df_05 = df[mask_05]

#Empiezo con 0.5

classstat = df_05["CLASSSTAT"]

mask_galaxies = (classstat < 0.5)

mask_stars = (classstat >= 0.5)

df_galaxies = df_05[mask_galaxies]

df_stars = df_05[mask_stars]

#Ya tengo diferenciadas galaxias y estrellas

#Ahora puedo empezar a usar los dataframes que diferencian galaxias y estrellas con las máscaras numeradas, y sacar lo que me interesa.

#El sufijo "g" lo uso de ahora en adelante para considerar galaxias, y el sufijo "s" para estrellas (stars)

mag_ks_g = df_galaxies["PETROMAG"]

mask_0_g = (mag_ks_g > 13.7)&(mag_ks_g <= 14.2)
mask_1_g = (mag_ks_g > 14.2)&(mag_ks_g <= 14.7)
mask_2_g = (mag_ks_g > 14.7)&(mag_ks_g <= 15.2)
mask_3_g = (mag_ks_g > 15.2)&(mag_ks_g <= 15.7)
mask_4_g = (mag_ks_g > 15.7)&(mag_ks_g <= 16.2)
mask_5_g = (mag_ks_g > 16.2)&(mag_ks_g <= 16.7)
mask_6_g = (mag_ks_g > 16.7)&(mag_ks_g <= 17.2)
mask_7_g = (mag_ks_g > 17.2)&(mag_ks_g <= 17.7)
mask_8_g = (mag_ks_g > 17.7)&(mag_ks_g <= 18.2)
mask_9_g = (mag_ks_g > 18.2)&(mag_ks_g <= 18.7)
mask_10_g = (mag_ks_g > 18.7)&(mag_ks_g <= 19.2)
mask_11_g = (mag_ks_g > 19.2)&(mag_ks_g <= 19.7)
mask_12_g = (mag_ks_g > 19.7)&(mag_ks_g <= 20.2)
mask_13_g = (mag_ks_g > 20.2)&(mag_ks_g <= 20.7)
mask_14_g = (mag_ks_g > 20.7)&(mag_ks_g <= 21.2)
mask_15_g = (mag_ks_g > 21.2)&(mag_ks_g <= 21.7)
mask_16_g = (mag_ks_g > 21.7)&(mag_ks_g <= 22.2)
mask_17_g = (mag_ks_g > 22.2)&(mag_ks_g <= 22.7)

df_galaxies_0 = df_galaxies[mask_0_g]
df_galaxies_1 = df_galaxies[mask_1_g]
df_galaxies_2 = df_galaxies[mask_2_g]
df_galaxies_3 = df_galaxies[mask_3_g]
df_galaxies_4 = df_galaxies[mask_4_g]
df_galaxies_5 = df_galaxies[mask_5_g]
df_galaxies_6 = df_galaxies[mask_6_g]
df_galaxies_7 = df_galaxies[mask_7_g]
df_galaxies_8 = df_galaxies[mask_8_g]
df_galaxies_9 = df_galaxies[mask_9_g]
df_galaxies_10 = df_galaxies[mask_10_g]
df_galaxies_11 = df_galaxies[mask_11_g]
df_galaxies_12 = df_galaxies[mask_12_g]
df_galaxies_13 = df_galaxies[mask_13_g]
df_galaxies_14 = df_galaxies[mask_14_g]
df_galaxies_15 = df_galaxies[mask_15_g]
df_galaxies_16 = df_galaxies[mask_16_g]
df_galaxies_17 = df_galaxies[mask_17_g]

mag_ks_s = df_stars["PETROMAG"]

mask_0_s = (mag_ks_s > 13.7)&(mag_ks_s <= 14.2)
mask_1_s = (mag_ks_s > 14.2)&(mag_ks_s <= 14.7)
mask_2_s = (mag_ks_s > 14.7)&(mag_ks_s <= 15.2)
mask_3_s = (mag_ks_s > 15.2)&(mag_ks_s <= 15.7)
mask_4_s = (mag_ks_s > 15.7)&(mag_ks_s <= 16.2)
mask_5_s = (mag_ks_s > 16.2)&(mag_ks_s <= 16.7)
mask_6_s = (mag_ks_s > 16.7)&(mag_ks_s <= 17.2)
mask_7_s = (mag_ks_s > 17.2)&(mag_ks_s <= 17.7)
mask_8_s = (mag_ks_s > 17.7)&(mag_ks_s <= 18.2)
mask_9_s = (mag_ks_s > 18.2)&(mag_ks_s <= 18.7)
mask_10_s = (mag_ks_s > 18.7)&(mag_ks_s <= 19.2)
mask_11_s = (mag_ks_s > 19.2)&(mag_ks_s <= 19.7)
mask_12_s = (mag_ks_s > 19.7)&(mag_ks_s <= 20.2)
mask_13_s = (mag_ks_s > 20.2)&(mag_ks_s <= 20.7)
mask_14_s = (mag_ks_s > 20.7)&(mag_ks_s <= 21.2)
mask_15_s = (mag_ks_s > 21.2)&(mag_ks_s <= 21.7)
mask_16_s = (mag_ks_s > 21.7)&(mag_ks_s <= 22.2)
mask_17_s = (mag_ks_s > 22.2)&(mag_ks_s <= 22.7)

df_stars_0 = df_stars[mask_0_s]
df_stars_1 = df_stars[mask_1_s]
df_stars_2 = df_stars[mask_2_s]
df_stars_3 = df_stars[mask_3_s]
df_stars_4 = df_stars[mask_4_s]
df_stars_5 = df_stars[mask_5_s]
df_stars_6 = df_stars[mask_6_s]
df_stars_7 = df_stars[mask_7_s]
df_stars_8 = df_stars[mask_8_s]
df_stars_9 = df_stars[mask_9_s]
df_stars_10 = df_stars[mask_10_s]
df_stars_11 = df_stars[mask_11_s]
df_stars_12 = df_stars[mask_12_s]
df_stars_13 = df_stars[mask_13_s]
df_stars_14 = df_stars[mask_14_s]
df_stars_15 = df_stars[mask_15_s]
df_stars_16 = df_stars[mask_16_s]
df_stars_17 = df_stars[mask_17_s]



#Ahora calculo el num de estrellas y galaxias, para cada rango.

galaxie_0 = len(df_galaxies_0)
galaxie_1 = len(df_galaxies_1)
galaxie_2 = len(df_galaxies_2)
galaxie_3 = len(df_galaxies_3)
galaxie_4 = len(df_galaxies_4)
galaxie_5 = len(df_galaxies_5)
galaxie_6 = len(df_galaxies_6)
galaxie_7 = len(df_galaxies_7)
galaxie_8 = len(df_galaxies_8)
galaxie_9 = len(df_galaxies_9)
galaxie_10 = len(df_galaxies_10)
galaxie_11 = len(df_galaxies_11)
galaxie_12 = len(df_galaxies_12)
galaxie_13 = len(df_galaxies_13)
galaxie_14 = len(df_galaxies_14)
galaxie_15 = len(df_galaxies_15)
galaxie_16 = len(df_galaxies_16)
galaxie_17 = len(df_galaxies_17)



galaxies_list_05 = np.array([galaxie_0, galaxie_1, galaxie_2, galaxie_3, galaxie_4, galaxie_5, galaxie_6, galaxie_7, galaxie_8, galaxie_9, galaxie_10, galaxie_11, galaxie_12, galaxie_13, galaxie_14, galaxie_15, galaxie_16, galaxie_17])



star_0 = len(df_stars_0)
star_1 = len(df_stars_1)
star_2 = len(df_stars_2)
star_3 = len(df_stars_3)
star_4 = len(df_stars_4)
star_5 = len(df_stars_5)
star_6 = len(df_stars_6)
star_7 = len(df_stars_7)
star_8 = len(df_stars_8)
star_9 = len(df_stars_9)
star_10 = len(df_stars_10)
star_11 = len(df_stars_11)
star_12 = len(df_stars_12)
star_13 = len(df_stars_13)
star_14 = len(df_stars_14)
star_15 = len(df_stars_15)
star_16 = len(df_stars_16)
star_17 = len(df_stars_17)


stars_list_05 = np.array([star_0, star_1, star_2, star_3, star_4, star_5, star_6, star_7, star_8, star_9, star_10, star_11, star_12, star_13, star_14, star_15, star_16, star_17])



#Sigo con 0.98


classstat = df_098["CLASSSTAT"]

mask_galaxies = (classstat < 0.975)

mask_stars = (classstat >= 0.975)

df_galaxies = df_098[mask_galaxies]

df_stars = df_098[mask_stars]

#Ya tengo diferenciadas galaxias y estrellas

#Ahora puedo empezar a usar los dataframes que diferencian galaxias y estrellas con las máscaras numeradas, y sacar lo que me interesa.

#El sufijo "g" lo uso de ahora en adelante para considerar galaxias, y el sufijo "s" para estrellas (stars)

mag_ks_g = df_galaxies["PETROMAG"]

mask_0_g = (mag_ks_g > 13.7)&(mag_ks_g <= 14.2)
mask_1_g = (mag_ks_g > 14.2)&(mag_ks_g <= 14.7)
mask_2_g = (mag_ks_g > 14.7)&(mag_ks_g <= 15.2)
mask_3_g = (mag_ks_g > 15.2)&(mag_ks_g <= 15.7)
mask_4_g = (mag_ks_g > 15.7)&(mag_ks_g <= 16.2)
mask_5_g = (mag_ks_g > 16.2)&(mag_ks_g <= 16.7)
mask_6_g = (mag_ks_g > 16.7)&(mag_ks_g <= 17.2)
mask_7_g = (mag_ks_g > 17.2)&(mag_ks_g <= 17.7)
mask_8_g = (mag_ks_g > 17.7)&(mag_ks_g <= 18.2)
mask_9_g = (mag_ks_g > 18.2)&(mag_ks_g <= 18.7)
mask_10_g = (mag_ks_g > 18.7)&(mag_ks_g <= 19.2)
mask_11_g = (mag_ks_g > 19.2)&(mag_ks_g <= 19.7)
mask_12_g = (mag_ks_g > 19.7)&(mag_ks_g <= 20.2)
mask_13_g = (mag_ks_g > 20.2)&(mag_ks_g <= 20.7)
mask_14_g = (mag_ks_g > 20.7)&(mag_ks_g <= 21.2)
mask_15_g = (mag_ks_g > 21.2)&(mag_ks_g <= 21.7)
mask_16_g = (mag_ks_g > 21.7)&(mag_ks_g <= 22.2)
mask_17_g = (mag_ks_g > 22.2)&(mag_ks_g <= 22.7)

df_galaxies_0 = df_galaxies[mask_0_g]
df_galaxies_1 = df_galaxies[mask_1_g]
df_galaxies_2 = df_galaxies[mask_2_g]
df_galaxies_3 = df_galaxies[mask_3_g]
df_galaxies_4 = df_galaxies[mask_4_g]
df_galaxies_5 = df_galaxies[mask_5_g]
df_galaxies_6 = df_galaxies[mask_6_g]
df_galaxies_7 = df_galaxies[mask_7_g]
df_galaxies_8 = df_galaxies[mask_8_g]
df_galaxies_9 = df_galaxies[mask_9_g]
df_galaxies_10 = df_galaxies[mask_10_g]
df_galaxies_11 = df_galaxies[mask_11_g]
df_galaxies_12 = df_galaxies[mask_12_g]
df_galaxies_13 = df_galaxies[mask_13_g]
df_galaxies_14 = df_galaxies[mask_14_g]
df_galaxies_15 = df_galaxies[mask_15_g]
df_galaxies_16 = df_galaxies[mask_16_g]
df_galaxies_17 = df_galaxies[mask_17_g]

mag_ks_s = df_stars["PETROMAG"]

mask_0_s = (mag_ks_s > 13.7)&(mag_ks_s <= 14.2)
mask_1_s = (mag_ks_s > 14.2)&(mag_ks_s <= 14.7)
mask_2_s = (mag_ks_s > 14.7)&(mag_ks_s <= 15.2)
mask_3_s = (mag_ks_s > 15.2)&(mag_ks_s <= 15.7)
mask_4_s = (mag_ks_s > 15.7)&(mag_ks_s <= 16.2)
mask_5_s = (mag_ks_s > 16.2)&(mag_ks_s <= 16.7)
mask_6_s = (mag_ks_s > 16.7)&(mag_ks_s <= 17.2)
mask_7_s = (mag_ks_s > 17.2)&(mag_ks_s <= 17.7)
mask_8_s = (mag_ks_s > 17.7)&(mag_ks_s <= 18.2)
mask_9_s = (mag_ks_s > 18.2)&(mag_ks_s <= 18.7)
mask_10_s = (mag_ks_s > 18.7)&(mag_ks_s <= 19.2)
mask_11_s = (mag_ks_s > 19.2)&(mag_ks_s <= 19.7)
mask_12_s = (mag_ks_s > 19.7)&(mag_ks_s <= 20.2)
mask_13_s = (mag_ks_s > 20.2)&(mag_ks_s <= 20.7)
mask_14_s = (mag_ks_s > 20.7)&(mag_ks_s <= 21.2)
mask_15_s = (mag_ks_s > 21.2)&(mag_ks_s <= 21.7)
mask_16_s = (mag_ks_s > 21.7)&(mag_ks_s <= 22.2)
mask_17_s = (mag_ks_s > 22.2)&(mag_ks_s <= 22.7)

df_stars_0 = df_stars[mask_0_s]
df_stars_1 = df_stars[mask_1_s]
df_stars_2 = df_stars[mask_2_s]
df_stars_3 = df_stars[mask_3_s]
df_stars_4 = df_stars[mask_4_s]
df_stars_5 = df_stars[mask_5_s]
df_stars_6 = df_stars[mask_6_s]
df_stars_7 = df_stars[mask_7_s]
df_stars_8 = df_stars[mask_8_s]
df_stars_9 = df_stars[mask_9_s]
df_stars_10 = df_stars[mask_10_s]
df_stars_11 = df_stars[mask_11_s]
df_stars_12 = df_stars[mask_12_s]
df_stars_13 = df_stars[mask_13_s]
df_stars_14 = df_stars[mask_14_s]
df_stars_15 = df_stars[mask_15_s]
df_stars_16 = df_stars[mask_16_s]
df_stars_17 = df_stars[mask_17_s]



#Ahora calculo el num de estrellas y galaxias, para cada rango.

galaxie_0 = len(df_galaxies_0)
galaxie_1 = len(df_galaxies_1)
galaxie_2 = len(df_galaxies_2)
galaxie_3 = len(df_galaxies_3)
galaxie_4 = len(df_galaxies_4)
galaxie_5 = len(df_galaxies_5)
galaxie_6 = len(df_galaxies_6)
galaxie_7 = len(df_galaxies_7)
galaxie_8 = len(df_galaxies_8)
galaxie_9 = len(df_galaxies_9)
galaxie_10 = len(df_galaxies_10)
galaxie_11 = len(df_galaxies_11)
galaxie_12 = len(df_galaxies_12)
galaxie_13 = len(df_galaxies_13)
galaxie_14 = len(df_galaxies_14)
galaxie_15 = len(df_galaxies_15)
galaxie_16 = len(df_galaxies_16)
galaxie_17 = len(df_galaxies_17)



galaxies_list_098 = np.array([galaxie_0, galaxie_1, galaxie_2, galaxie_3, galaxie_4, galaxie_5, galaxie_6, galaxie_7, galaxie_8, galaxie_9, galaxie_10, galaxie_11, galaxie_12, galaxie_13, galaxie_14, galaxie_15, galaxie_16, galaxie_17])



star_0 = len(df_stars_0)
star_1 = len(df_stars_1)
star_2 = len(df_stars_2)
star_3 = len(df_stars_3)
star_4 = len(df_stars_4)
star_5 = len(df_stars_5)
star_6 = len(df_stars_6)
star_7 = len(df_stars_7)
star_8 = len(df_stars_8)
star_9 = len(df_stars_9)
star_10 = len(df_stars_10)
star_11 = len(df_stars_11)
star_12 = len(df_stars_12)
star_13 = len(df_stars_13)
star_14 = len(df_stars_14)
star_15 = len(df_stars_15)
star_16 = len(df_stars_16)
star_17 = len(df_stars_17)


stars_list_098 = np.array([star_0, star_1, star_2, star_3, star_4, star_5, star_6, star_7, star_8, star_9, star_10, star_11, star_12, star_13, star_14, star_15, star_16, star_17])


stars_total = stars_list_05 + stars_list_098
galaxies_total = galaxies_list_05 + galaxies_list_098

area_sharks = 7.23 #revisar, este dato no lo recuero y perdí el correo que me decía Aurelio, es en grados de arco

N_stars = stars_total/area_sharks
N_galaxies = galaxies_total/area_sharks

area_sharks_mins = 7.23*3600 #revisar, este dato no lo recuero y perdí el correo que me decía Aurelio, es en grados de arco

N_stars_mins = stars_total/area_sharks_mins
N_galaxies_mins = galaxies_total/area_sharks_mins


ks_mitad_intervalos = ([14.2, 14.7, 15.2, 15.7, 16.2, 16.7, 17.2, 17.7, 18.2, 18.7, 19.2, 19.7, 20.2, 20.7, 21.2, 21.7, 22.2, 22.7])

fig = plt.figure()
plt.plot(ks_mitad_intervalos,np.log(N_stars), "*" ,label="Stars in Sharks")
plt.plot(ks_mitad_intervalos,np.log(N_galaxies), "." ,label="Galaxies in Sharks")
plt.xlabel("Ks")
plt.ylabel(r"$log N ~ (objects/deg^2/ 0.5 mag)$")
#fig.xaxis.set_major_locator(ticker.MultipleLocator(0.5))
#fig.xaxis.set_minor_locator(ticker.MultipleLocator(0.1))
plt.xticks([13, 15, 17, 19, 21, 23 ])
plt.legend()

plt.savefig("Stars_galaxies_eros.png")





#Ploteo también distribución espacial de EROs. El dataframe de todos los eros se llama simplemente df

plt.figure()
plt.plot(df["RA"], df["DEC"], ".", markersize = 1.5)
plt.xlabel("RA")
plt.ylabel("DEC")
plt.savefig("EROs_distribution_RA_DEC.png")










wb = Workbook()
ruta = 'salida_EROS.xlsx'

hoja1 = wb.active
hoja1.title = "hoja"
fila = 1 #Fila donde empezamos
col_stars = 1 #Columna donde guardo los valores
col_galaxies = 2
col_n_stars = 3
col_n_galaxies = 4
col_n_min_stars = 5
col_n_min_galaxies = 6

for stars, galaxies, n_stars, n_galaxies, n_min_stars, n_min_galaxies in zip(stars_total, galaxies_total, N_stars, N_galaxies, N_stars_mins, N_galaxies_mins):  #Aquí pongo en "in" el valor que quiero sacar como columna
    hoja1.cell(column=col_stars, row=fila, value=stars)
    hoja1.cell(column=col_galaxies, row=fila, value=galaxies)
    hoja1.cell(column=col_n_stars, row=fila, value=n_stars)
    hoja1.cell(column=col_n_galaxies, row=fila, value=n_galaxies)
    hoja1.cell(column=col_n_min_stars, row=fila, value=n_min_stars)
    hoja1.cell(column=col_n_min_galaxies, row=fila, value=n_min_galaxies)
    fila+=1
    


wb.save(filename = ruta)


























