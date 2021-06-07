# -*- coding: utf-8 -*-
"""
Created on Sat Jun  5 20:52:58 2021

@author: Sergi
"""


import numpy as np

import astropy as astropy 

import pandas as pd

import seaborn as seaborn

import matplotlib.pyplot as plt

from astropy.table import Table, vstack    #Ojo importante aquí importar "Table" con mayuscula

from openpyxl import Workbook

#Leo datos (Objetos comunes de SHARKS y DES)
 

df=Table.read("fits/sharks_sgpe_nodif_signoise5_ALL_EROs.fits", format="fits")  #Obj en sharks y des, con 5sigma

mag_ks = df["PETROMAG"]

mask_05 = (mag_ks<19.2)#&(mag_des_r <= 24.65)  #Si aplico las dos condiciones, se limita el numero de objetos (pierdo objetos)
mask_098 = (mag_ks>=19.2)#&(mag_des_r > 24.65)

df_098 = df[mask_098]
df_05 = df[mask_05]

#Empiezo con 0.5

classstat = df_05["CLASSSTAT"]

mask_galaxies = (classstat < 0.5)

mask_stars = (classstat >= 0.5)

df_galaxies = df_05[mask_galaxies]

df_stars = df_05[mask_stars]

#Ya tengo diferenciadas galaxias y estrellas

#Ahora puedo empezar a usar los dataframes que diferencian galaxias y estrellas con las máscaras numeradas, y sacar lo que me interesa.

#El sufijo "g" lo uso de ahora en adelante para considerar galaxias, y el sufijo "s" para estrellas (stars)

mag_ks_g = df_galaxies["PETROMAG"]

mask_00_g = (mag_ks_g <= 13.7)
mask_0_g = (mag_ks_g > 13.7)&(mag_ks_g <= 14.2)
mask_1_g = (mag_ks_g > 14.2)&(mag_ks_g <= 14.7)
mask_2_g = (mag_ks_g > 14.7)&(mag_ks_g <= 15.2)
mask_3_g = (mag_ks_g > 15.2)&(mag_ks_g <= 15.7)
mask_4_g = (mag_ks_g > 15.7)&(mag_ks_g <= 16.2)
mask_5_g = (mag_ks_g > 16.2)&(mag_ks_g <= 16.7)
mask_6_g = (mag_ks_g > 16.7)&(mag_ks_g <= 17.2)
mask_7_g = (mag_ks_g > 17.2)&(mag_ks_g <= 17.7)
mask_8_g = (mag_ks_g > 17.7)&(mag_ks_g <= 18.2)
mask_9_g = (mag_ks_g > 18.2)&(mag_ks_g <= 18.7)
mask_10_g = (mag_ks_g > 18.7)&(mag_ks_g <= 19.2)
mask_11_g = (mag_ks_g > 19.2)&(mag_ks_g <= 19.7)
mask_12_g = (mag_ks_g > 19.7)&(mag_ks_g <= 20.2)
mask_13_g = (mag_ks_g > 20.2)&(mag_ks_g <= 20.7)
mask_14_g = (mag_ks_g > 20.7)&(mag_ks_g <= 21.2)
mask_15_g = (mag_ks_g > 21.2)&(mag_ks_g <= 21.7)
mask_16_g = (mag_ks_g > 21.7)&(mag_ks_g <= 22.2)
mask_17_g = (mag_ks_g > 22.2)&(mag_ks_g <= 22.7)
mask_18_g = (mag_ks_g > 22.7)

df_galaxies_00_05 = df_galaxies[mask_00_g]
df_galaxies_0_05 = df_galaxies[mask_0_g]
df_galaxies_1_05 = df_galaxies[mask_1_g]
df_galaxies_2_05 = df_galaxies[mask_2_g]
df_galaxies_3_05 = df_galaxies[mask_3_g]
df_galaxies_4_05 = df_galaxies[mask_4_g]
df_galaxies_5_05 = df_galaxies[mask_5_g]
df_galaxies_6_05 = df_galaxies[mask_6_g]
df_galaxies_7_05 = df_galaxies[mask_7_g]
df_galaxies_8_05 = df_galaxies[mask_8_g]
df_galaxies_9_05 = df_galaxies[mask_9_g]
df_galaxies_10_05 = df_galaxies[mask_10_g]
df_galaxies_11_05 = df_galaxies[mask_11_g]
df_galaxies_12_05 = df_galaxies[mask_12_g]
df_galaxies_13_05 = df_galaxies[mask_13_g]
df_galaxies_14_05 = df_galaxies[mask_14_g]
df_galaxies_15_05 = df_galaxies[mask_15_g]
df_galaxies_16_05 = df_galaxies[mask_16_g]
df_galaxies_17_05 = df_galaxies[mask_17_g]
df_galaxies_18_05 = df_galaxies[mask_18_g]

#Calculo los R-Ks de cada sección 





mag_ks_s = df_stars["PETROMAG"]

mask_00_s = (mag_ks_s <= 13.7)
mask_0_s = (mag_ks_s > 13.7)&(mag_ks_s <= 14.2)
mask_1_s = (mag_ks_s > 14.2)&(mag_ks_s <= 14.7)
mask_2_s = (mag_ks_s > 14.7)&(mag_ks_s <= 15.2)
mask_3_s = (mag_ks_s > 15.2)&(mag_ks_s <= 15.7)
mask_4_s = (mag_ks_s > 15.7)&(mag_ks_s <= 16.2)
mask_5_s = (mag_ks_s > 16.2)&(mag_ks_s <= 16.7)
mask_6_s = (mag_ks_s > 16.7)&(mag_ks_s <= 17.2)
mask_7_s = (mag_ks_s > 17.2)&(mag_ks_s <= 17.7)
mask_8_s = (mag_ks_s > 17.7)&(mag_ks_s <= 18.2)
mask_9_s = (mag_ks_s > 18.2)&(mag_ks_s <= 18.7)
mask_10_s = (mag_ks_s > 18.7)&(mag_ks_s <= 19.2)
mask_11_s = (mag_ks_s > 19.2)&(mag_ks_s <= 19.7)
mask_12_s = (mag_ks_s > 19.7)&(mag_ks_s <= 20.2)
mask_13_s = (mag_ks_s > 20.2)&(mag_ks_s <= 20.7)
mask_14_s = (mag_ks_s > 20.7)&(mag_ks_s <= 21.2)
mask_15_s = (mag_ks_s > 21.2)&(mag_ks_s <= 21.7)
mask_16_s = (mag_ks_s > 21.7)&(mag_ks_s <= 22.2)
mask_17_s = (mag_ks_s > 22.2)&(mag_ks_s <= 22.7)
mask_18_s = (mag_ks_s > 22.7)

df_stars_00_05 = df_stars[mask_00_s]
df_stars_0_05 = df_stars[mask_0_s]
df_stars_1_05 = df_stars[mask_1_s]
df_stars_2_05 = df_stars[mask_2_s]
df_stars_3_05 = df_stars[mask_3_s]
df_stars_4_05 = df_stars[mask_4_s]
df_stars_5_05 = df_stars[mask_5_s]
df_stars_6_05 = df_stars[mask_6_s]
df_stars_7_05 = df_stars[mask_7_s]
df_stars_8_05 = df_stars[mask_8_s]
df_stars_9_05 = df_stars[mask_9_s]
df_stars_10_05 = df_stars[mask_10_s]
df_stars_11_05 = df_stars[mask_11_s]
df_stars_12_05 = df_stars[mask_12_s]
df_stars_13_05 = df_stars[mask_13_s]
df_stars_14_05 = df_stars[mask_14_s]
df_stars_15_05 = df_stars[mask_15_s]
df_stars_16_05 = df_stars[mask_16_s]
df_stars_17_05 = df_stars[mask_17_s]
df_stars_18_05 = df_stars[mask_18_s]


#Sigo con 0.98


classstat = df_098["CLASSSTAT"]

mask_galaxies = (classstat < 0.975)

mask_stars = (classstat >= 0.975)

df_galaxies = df_098[mask_galaxies]

df_stars = df_098[mask_stars]

#Ya tengo diferenciadas galaxias y estrellas

#Ahora puedo empezar a usar los dataframes que diferencian galaxias y estrellas con las máscaras numeradas, y sacar lo que me interesa.

#El sufijo "g" lo uso de ahora en adelante para considerar galaxias, y el sufijo "s" para estrellas (stars)

mag_ks_g = df_galaxies["PETROMAG"]

mask_00_g = (mag_ks_g <= 13.7)
mask_0_g = (mag_ks_g > 13.7)&(mag_ks_g <= 14.2)
mask_1_g = (mag_ks_g > 14.2)&(mag_ks_g <= 14.7)
mask_2_g = (mag_ks_g > 14.7)&(mag_ks_g <= 15.2)
mask_3_g = (mag_ks_g > 15.2)&(mag_ks_g <= 15.7)
mask_4_g = (mag_ks_g > 15.7)&(mag_ks_g <= 16.2)
mask_5_g = (mag_ks_g > 16.2)&(mag_ks_g <= 16.7)
mask_6_g = (mag_ks_g > 16.7)&(mag_ks_g <= 17.2)
mask_7_g = (mag_ks_g > 17.2)&(mag_ks_g <= 17.7)
mask_8_g = (mag_ks_g > 17.7)&(mag_ks_g <= 18.2)
mask_9_g = (mag_ks_g > 18.2)&(mag_ks_g <= 18.7)
mask_10_g = (mag_ks_g > 18.7)&(mag_ks_g <= 19.2)
mask_11_g = (mag_ks_g > 19.2)&(mag_ks_g <= 19.7)
mask_12_g = (mag_ks_g > 19.7)&(mag_ks_g <= 20.2)
mask_13_g = (mag_ks_g > 20.2)&(mag_ks_g <= 20.7)
mask_14_g = (mag_ks_g > 20.7)&(mag_ks_g <= 21.2)
mask_15_g = (mag_ks_g > 21.2)&(mag_ks_g <= 21.7)
mask_16_g = (mag_ks_g > 21.7)&(mag_ks_g <= 22.2)
mask_17_g = (mag_ks_g > 22.2)&(mag_ks_g <= 22.7)
mask_18_g = (mag_ks_g > 22.7)

df_galaxies_00_098 = df_galaxies[mask_00_g]
df_galaxies_0_098 = df_galaxies[mask_0_g]
df_galaxies_1_098 = df_galaxies[mask_1_g]
df_galaxies_2_098 = df_galaxies[mask_2_g]
df_galaxies_3_098 = df_galaxies[mask_3_g]
df_galaxies_4_098 = df_galaxies[mask_4_g]
df_galaxies_5_098 = df_galaxies[mask_5_g]
df_galaxies_6_098 = df_galaxies[mask_6_g]
df_galaxies_7_098 = df_galaxies[mask_7_g]
df_galaxies_8_098 = df_galaxies[mask_8_g]
df_galaxies_9_098 = df_galaxies[mask_9_g]
df_galaxies_10_098 = df_galaxies[mask_10_g]
df_galaxies_11_098 = df_galaxies[mask_11_g]
df_galaxies_12_098 = df_galaxies[mask_12_g]
df_galaxies_13_098 = df_galaxies[mask_13_g]
df_galaxies_14_098 = df_galaxies[mask_14_g]
df_galaxies_15_098 = df_galaxies[mask_15_g]
df_galaxies_16_098 = df_galaxies[mask_16_g]
df_galaxies_17_098 = df_galaxies[mask_17_g]
df_galaxies_18_098 = df_galaxies[mask_18_g]








mag_ks_s = df_stars["PETROMAG"]

mask_00_s = (mag_ks_s <= 13.7)
mask_0_s = (mag_ks_s > 13.7)&(mag_ks_s <= 14.2)
mask_1_s = (mag_ks_s > 14.2)&(mag_ks_s <= 14.7)
mask_2_s = (mag_ks_s > 14.7)&(mag_ks_s <= 15.2)
mask_3_s = (mag_ks_s > 15.2)&(mag_ks_s <= 15.7)
mask_4_s = (mag_ks_s > 15.7)&(mag_ks_s <= 16.2)
mask_5_s = (mag_ks_s > 16.2)&(mag_ks_s <= 16.7)
mask_6_s = (mag_ks_s > 16.7)&(mag_ks_s <= 17.2)
mask_7_s = (mag_ks_s > 17.2)&(mag_ks_s <= 17.7)
mask_8_s = (mag_ks_s > 17.7)&(mag_ks_s <= 18.2)
mask_9_s = (mag_ks_s > 18.2)&(mag_ks_s <= 18.7)
mask_10_s = (mag_ks_s > 18.7)&(mag_ks_s <= 19.2)
mask_11_s = (mag_ks_s > 19.2)&(mag_ks_s <= 19.7)
mask_12_s = (mag_ks_s > 19.7)&(mag_ks_s <= 20.2)
mask_13_s = (mag_ks_s > 20.2)&(mag_ks_s <= 20.7)
mask_14_s = (mag_ks_s > 20.7)&(mag_ks_s <= 21.2)
mask_15_s = (mag_ks_s > 21.2)&(mag_ks_s <= 21.7)
mask_16_s = (mag_ks_s > 21.7)&(mag_ks_s <= 22.2)
mask_17_s = (mag_ks_s > 22.2)&(mag_ks_s <= 22.7)
mask_18_s = (mag_ks_s > 22.7)

df_stars_00_098 = df_stars[mask_00_s]
df_stars_0_098 = df_stars[mask_0_s]
df_stars_1_098 = df_stars[mask_1_s]
df_stars_2_098 = df_stars[mask_2_s]
df_stars_3_098 = df_stars[mask_3_s]
df_stars_4_098 = df_stars[mask_4_s]
df_stars_5_098 = df_stars[mask_5_s]
df_stars_6_098 = df_stars[mask_6_s]
df_stars_7_098 = df_stars[mask_7_s]
df_stars_8_098 = df_stars[mask_8_s]
df_stars_9_098 = df_stars[mask_9_s]
df_stars_10_098 = df_stars[mask_10_s]
df_stars_11_098 = df_stars[mask_11_s]
df_stars_12_098 = df_stars[mask_12_s]
df_stars_13_098 = df_stars[mask_13_s]
df_stars_14_098 = df_stars[mask_14_s]
df_stars_15_098 = df_stars[mask_15_s]
df_stars_16_098 = df_stars[mask_16_s]
df_stars_17_098 = df_stars[mask_17_s]
df_stars_18_098 = df_stars[mask_18_s]





#Junto todas las estrellas
df_stars_00 = vstack([df_stars_00_05, df_stars_00_098])
df_stars_0 = vstack([df_stars_0_05, df_stars_0_098])
df_stars_1 = vstack([df_stars_1_05, df_stars_1_098])
df_stars_2 = vstack([df_stars_2_05, df_stars_2_098])
df_stars_3 = vstack([df_stars_3_05, df_stars_3_098])
df_stars_4 = vstack([df_stars_4_05, df_stars_4_098])
df_stars_5 = vstack([df_stars_5_05, df_stars_5_098])
df_stars_6 = vstack([df_stars_6_05, df_stars_6_098])
df_stars_7 = vstack([df_stars_7_05, df_stars_7_098])
df_stars_8 = vstack([df_stars_8_05, df_stars_8_098])
df_stars_9 = vstack([df_stars_9_05, df_stars_9_098])
df_stars_10 = vstack([df_stars_10_05, df_stars_10_098])
df_stars_11 = vstack([df_stars_11_05, df_stars_11_098])
df_stars_12 = vstack([df_stars_12_05, df_stars_12_098])
df_stars_13 = vstack([df_stars_13_05, df_stars_13_098])
df_stars_14 = vstack([df_stars_14_05, df_stars_14_098])
df_stars_15 = vstack([df_stars_15_05, df_stars_15_098])
df_stars_16 = vstack([df_stars_16_05, df_stars_16_098])
df_stars_17 = vstack([df_stars_17_05, df_stars_17_098])
df_stars_18 = vstack([df_stars_18_05, df_stars_18_098])

df_stars_all = vstack([df_stars_00,df_stars_0, df_stars_1, df_stars_2, df_stars_3, df_stars_4, df_stars_5, df_stars_6, df_stars_7, df_stars_8, df_stars_9, df_stars_10, df_stars_11, df_stars_12, df_stars_13, df_stars_14, df_stars_15, df_stars_16, df_stars_17,  df_stars_18])

stars_list = np.array([len(df_stars_00), len(df_stars_0), len(df_stars_1), len(df_stars_2), len(df_stars_3), len(df_stars_4), len(df_stars_5), len(df_stars_6), len(df_stars_7), len(df_stars_8), len(df_stars_9), len(df_stars_10), len(df_stars_11), len(df_stars_12), len(df_stars_13), len(df_stars_14),  len(df_stars_15), len(df_stars_16), len(df_stars_17),len(df_stars_18)])
total_stars = len(df_stars_all)



#Junto todas las galaxias
df_galaxies_00 = vstack([df_galaxies_00_05, df_galaxies_00_098])
df_galaxies_0 = vstack([df_galaxies_0_05, df_galaxies_0_098])
df_galaxies_1 = vstack([df_galaxies_1_05, df_galaxies_1_098])
df_galaxies_2 = vstack([df_galaxies_2_05, df_galaxies_2_098])
df_galaxies_3 = vstack([df_galaxies_3_05, df_galaxies_3_098])
df_galaxies_4 = vstack([df_galaxies_4_05, df_galaxies_4_098])
df_galaxies_5 = vstack([df_galaxies_5_05, df_galaxies_5_098])
df_galaxies_6 = vstack([df_galaxies_6_05, df_galaxies_6_098])
df_galaxies_7 = vstack([df_galaxies_7_05, df_galaxies_7_098])
df_galaxies_8 = vstack([df_galaxies_8_05, df_galaxies_8_098])
df_galaxies_9 = vstack([df_galaxies_9_05, df_galaxies_9_098])
df_galaxies_10 = vstack([df_galaxies_10_05, df_galaxies_10_098])
df_galaxies_11 = vstack([df_galaxies_11_05, df_galaxies_11_098])
df_galaxies_12 = vstack([df_galaxies_12_05, df_galaxies_12_098])
df_galaxies_13 = vstack([df_galaxies_13_05, df_galaxies_13_098])
df_galaxies_14 = vstack([df_galaxies_14_05, df_galaxies_14_098])
df_galaxies_15 = vstack([df_galaxies_15_05, df_galaxies_15_098])
df_galaxies_16 = vstack([df_galaxies_16_05, df_galaxies_16_098])
df_galaxies_17 = vstack([df_galaxies_17_05, df_galaxies_17_098])
df_galaxies_18 = vstack([df_galaxies_18_05, df_galaxies_18_098])

df_galaxies_all = vstack([df_galaxies_00, df_galaxies_0, df_galaxies_1, df_galaxies_2, df_galaxies_3, df_galaxies_4, df_galaxies_5, df_galaxies_6, df_galaxies_7, df_galaxies_8, df_galaxies_9, df_galaxies_10, df_galaxies_11, df_galaxies_12, df_galaxies_13, df_galaxies_14, df_galaxies_15, df_galaxies_16, df_galaxies_17,  df_galaxies_18])

galaxies_list = np.array([len(df_galaxies_00), len(df_galaxies_0), len(df_galaxies_1), len(df_galaxies_2), len(df_galaxies_3), len(df_galaxies_4), len(df_galaxies_5), len(df_galaxies_6), len(df_galaxies_7), len(df_galaxies_8), len(df_galaxies_9), len(df_galaxies_10), len(df_galaxies_11), len(df_galaxies_12), len(df_galaxies_13), len(df_galaxies_14),  len(df_galaxies_15), len(df_galaxies_16), len(df_galaxies_17), len(df_galaxies_18)])
total_galaxies = len(df_galaxies_all)

print(stars_list)
print(galaxies_list)





#Retiro último valor de estrellas, que es cero, para evitar problemas

stars_list_sin_0 = stars_list[0:17]

#Calculo "N", cuentas/deg^2

area_sharks = 7.23 


N_stars = stars_list_sin_0/area_sharks
N_galaxies = galaxies_list/area_sharks





#Creo mis ejes x, y los ejes de Daddi.

ks_sharks_vega = [12, 12.5, 13, 13.5, 14, 14.5, 15, 15.5, 16, 16.5, 17, 17.5, 18, 18.5, 19, 19.5, 20, 20.5]
ks_sharks_AB = []
for i in range(len(ks_sharks_vega)):
    ks_sharks_AB.append(ks_sharks_vega[i]+1.83)

#Para Ks de estrellas quito el último valor(porque antes quité el último de estrellas)
ks_sharks_AB_stars = ks_sharks_AB[0:17]


#Calculo ahora el N de Daddi (con los datos del artículo)


ks_Daddi_vega = [12, 12.5, 13, 13.5, 14, 14.5, 15, 15.5, 16, 16.5, 17, 17.5, 18, 18.5, 19]
ks_Daddi_AB = []
for i in range(len(ks_Daddi_vega)):
    ks_Daddi_AB.append(ks_Daddi_vega[i]+1.83)
    


"""
fig = plt.figure()
plt.plot(ks_mitad_intervalos,np.log(N_stars), "*" ,label="Stars in Sharks")
plt.plot(ks_mitad_intervalos,np.log(N_galaxies), "." ,label="Galaxies in Sharks")
plt.xlabel("Ks")
plt.ylabel(r"$log N ~ (objects/deg^2/ 0.5 mag)$")
#fig.xaxis.set_major_locator(ticker.MultipleLocator(0.5))
#fig.xaxis.set_minor_locator(ticker.MultipleLocator(0.1))
plt.xticks([13, 15, 17, 19, 21, 23 ])
plt.legend()

plt.savefig("Stars_galaxies_eros.png")
"""

imagen1 = seaborn.jointplot(df["RA"], df["DEC"], kind="hex")
imagen1.fig.suptitle("EROs distribution", fontsize=11)
imagen1.set_axis_labels(r'RA $[deg]$', 'DEC $[deg]$', fontsize=11)
imagen1.fig.axes[0].invert_xaxis()

imagen2 = seaborn.jointplot(df_galaxies_all["RA"], df_galaxies_all["DEC"], kind="hex")
imagen2.fig.suptitle("EROs galaxies distribution", fontsize=11)
imagen2.set_axis_labels(r'RA $[deg]$', 'DEC $[deg]$', fontsize=11)
imagen2.fig.axes[0].invert_xaxis()
#Ploteo también distribución espacial de EROs. El dataframe de todos los eros se llama simplemente df

plt.figure()
plt.plot(df["RA"], df["DEC"], ",", markersize = 1)
plt.xlabel("RA [$deg$]")
plt.ylabel("Dec [$deg$]")
plt.title("EROs distribution")
plt.gca().invert_xaxis()
plt.show()
plt.savefig("EROs_distribution_RA_DEC.png")


#Ploteo la distribución de galaxias y estrellas por separado

fig = plt.figure()
plt.plot(df_galaxies_all["RA"], df_galaxies_all["DEC"], ",", markersize = 1, label = "EROs detected as galaxies", color = "grey", alpha =0.8)
plt.plot(df_stars_all["RA"], df_stars_all["DEC"], ",", markersize = 1, label = "EROs detected as stars", color = "red", alpha =1)
plt.xlabel("RA [$deg$]")
plt.ylabel("Dec [$deg$]")
plt.title("EROs galaxies/stars distribution")
plt.gca().invert_xaxis()
plt.show()
plt.legend()
plt.savefig("EROs_galax_stars_distribution_RA_DEC.png")

"""
wb = Workbook()
ruta = 'salida_EROS.xlsx'

hoja1 = wb.active
hoja1.title = "hoja"
fila = 1 #Fila donde empezamos
col_stars = 1 #Columna donde guardo los valores
col_galaxies = 2
col_n_stars = 3
col_n_galaxies = 4


for stars, galaxies, n_stars, n_galaxies in zip(stars_list, galaxies_list, N_stars, N_galaxies):
    hoja1.cell(column=col_stars, row=fila, value=stars)
    hoja1.cell(column=col_galaxies, row=fila, value=galaxies)
    hoja1.cell(column=col_n_stars, row=fila, value=n_stars)
    hoja1.cell(column=col_n_galaxies, row=fila, value=n_galaxies)
    fila+=1
    


wb.save(filename = ruta)
"""